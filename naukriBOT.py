# -*- coding: utf-8 -*-


from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
import openpyxl,time
import pandas as pd
import lxml
import html5lib, os
from pandas.io.html import read_html


#Provie Chrome driver path
driver = webdriver.Chrome(executable_path= r'C:\Users\XXXXXXX\python-workspace\chromedriver.exe')
driver.maximize_window()

#Website address
driver.get("https://www.naukri.com/")

#Counter
count = 0

#excel part
os.chdir("C:\\Users\\XXXXXXX\\python-workspace")
wb = openpyxl.load_workbook("nBotFile.xlsx")
sheet = wb["Sheet1"]
rowCount = sheet.max_row # get the number of rows present in sheet
colCount = sheet.max_column # get the number of columns in sheet
a=[]




driver.switch_to_window(driver.window_handles[1])
driver.close()
driver.switch_to_window(driver.window_handles[1])
driver.close()
driver.switch_to_window(driver.window_handles[0])

login = WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH,'//*[@id="login_Layer"]/div')))
login.click()

#UserName
userName = WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH,'//*[@id="root"]/div[2]/div[2]/div/form/div[2]/input')))
userName.send_keys("provide your username")

#password
password = WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH,'//*[@id="root"]/div[2]/div[2]/div/form/div[3]/input')))
password.send_keys("provide your naukri password here")

#Login Button
Button = WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH,'//*[@id="root"]/div[2]/div[2]/div/form/div[6]/button')))
Button.click()

#Click on View All
viewAll = WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH,"//section[@class='combine-margins recoJobs']/div[2]/div/div[3]/a")))
viewAll.click()

#List of all jobs
time.sleep(5)
listofJobs = driver.find_elements_by_xpath('//*[@id="root"]/div[2]/div[1]/div[1]/div[1]/div[3]/article')
print("Lenght of Job List: ",len(listofJobs))


for i in listofJobs:
    try:
        if count < 5:
            i.click()
            driver.switch_to_window(driver.window_handles[1])
            apply = WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH,'//*[@id="root"]/main/div[2]/div[2]/section[1]/div[1]/div[3]/div/button[2]')))
            print(apply.text)
            if apply.text == "APPLY ON COMPANY SITE":
                print("Automation will not go to Company Site")
            else:
                apply.click()
                count=count+1
            time.sleep(5)
            driver.close()
            time.sleep(1)
            driver.switch_to_window(driver.window_handles[0])
        else:
            break
        
    except:
        print("i am on except")
        driver.close()
        driver.switch_to_window(driver.window_handles[0])   


# Update resume headline

myNaukri = WebDriverWait(driver,20).until(EC.visibility_of_element_located((By.XPATH,"//div[contains(@class,'mTxt') and text()='My Naukri']")))
action = ActionChains(driver);
action.move_to_element(myNaukri).perform();
myEdit = WebDriverWait(driver,20).until(EC.visibility_of_element_located((By.XPATH,'//*[@id="root"]/header/div/div/ul[2]/li[2]/div/ul[1]/li[1]/a')))
action.move_to_element(myEdit).perform();
myEdit.click()



# headline
headline = WebDriverWait(driver,20).until(EC.visibility_of_element_located((By.XPATH,'//*[@id="lazyResumeHead"]/div/div/div/div[1]/span[2]')))
headline.click()

changeHeadline = WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="resumeHeadlineTxt"]')))
changeHeadline.send_keys(Keys.CONTROL, 'a')
changeHeadline.send_keys(Keys.CONTROL, 'c')
changeHeadline.send_keys(Keys.CONTROL, 'p')

#click on save
save = WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH, "//button[contains(@class,'waves-effect waves-light btn-large blue-btn') and text()='Save']")))
save.click()










